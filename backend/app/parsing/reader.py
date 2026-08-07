"""Reads a client workbook into raw positional rows.

Three defences, all against traps verified in the real files
(docs/FINDINGS.md §2):

1. **Sheet selection is by header signature**, never by name or index. The
   samples use `sheet1`, `Sheet1` and `Sheet2` for equivalent content, and at
   least one workbook opens on a decorative cover page.
2. **Reported dimensions are ignored.** Real sheets claim ~1,048,570 rows and
   hold a few hundred; a naive loop processes a million empty rows. Reading
   stops after a run of consecutive blanks.
3. **Junk columns are kept, not dropped.** The Apps Tracker's fourth column has
   a single space for a header but carries a real 2-character recruiter code.
   It is excluded from exports but preserved in the raw row.

Cells are read twice — once for cached values, once for formulas — so a
formula that Excel already computed keeps its answer, and one that it did not
still reaches the normalizers as a string to evaluate.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Stop after this many consecutive blank rows. Comfortably larger than any
# formatting gap seen in the samples, small enough that a phantom sheet costs
# nothing to reject.
BLANK_RUN_LIMIT = 50

# Never scan more than this, whatever the file claims. A safety net for a
# pathological upload; the blank-run rule ends normal files long before.
MAX_ROWS_SCANNED = 250_000

# How many rows below the top to search for the header row. Some exports carry
# a title block above the real headers.
HEADER_SEARCH_DEPTH = 10


class NoDataSheetError(Exception):
    """No worksheet in the workbook matched a known header signature."""


@dataclass(frozen=True)
class HeaderSignature:
    """Identifies a kind of file by the headers it must contain."""

    name: str
    required: frozenset[str]
    minimum_matches: int

    def matches(self, headers: list[str]) -> bool:
        normalized = {_canonical(h) for h in headers if h}
        return len(self.required & normalized) >= self.minimum_matches


def _canonical(header: object) -> str:
    """Case- and spacing-insensitive header key, so drift does not break matching."""
    return re.sub(r"\s+", " ", str(header)).strip().casefold()


# Signatures use columns UNIQUE to each file. The two schemas overlap on
# SERIAL NO, STATUS DATE, REASON, CHQ/MO/PO, CREDIT CARD and ACCOUNT NUMBER, so
# matching on shared columns would make either file match either signature.
STATUS_REPORT = HeaderSignature(
    name="status_report",
    required=frozenset(
        _canonical(h)
        for h in (
            "STATUS DESCRIPTION",
            "REASONDESC",
            "SG BATCH NO",
            "A0 Attempts",
            "Recruiter Batch No",
            "DonationAmount",
            "DEBIT_CREDIT_CARD",
            "LocationCode",
            "Recruiter Code",
            "CUSTOMERS NAME",
        )
    ),
    minimum_matches=5,
)

APPS_TRACKER = HeaderSignature(
    name="apps_tracker",
    required=frozenset(
        _canonical(h)
        for h in (
            "IMPORTANT REMARKS",
            "SUB-RECRUITER CODE",
            "PROFILE TYPE",
            "CUSTOMER'S NAME",
            "VERIFIEDDATE",
            "DEBIT DATE",
            "Fundraiser Name",
            "CANCELLATION DATE",
            "OTHER NOTES",
            "DONATION AMOUNT",
            "EVENT CODE",
            "SIGNUP DATE",
        )
    ),
    minimum_matches=5,
)

SIGNATURES = (APPS_TRACKER, STATUS_REPORT)


def detect_signature(headers: list[str]) -> HeaderSignature | None:
    for signature in SIGNATURES:
        if signature.matches(headers):
            return signature
    return None


@dataclass(frozen=True)
class RawRow:
    """One spreadsheet row, positionally intact.

    Positional rather than keyed by header because headers are not unique —
    the Apps Tracker has two junk columns whose headers are `' '` and `''`, and
    keying by name would silently collapse them.
    """

    number: int
    """1-based row number in the sheet, so an exception can point a human at it."""

    cells: tuple[Any, ...]
    headers: tuple[str, ...]

    def get(self, header: str, default: Any = None) -> Any:
        wanted = _canonical(header)
        for index, name in enumerate(self.headers):
            if _canonical(name) == wanted:
                return self.cells[index] if index < len(self.cells) else default
        return default

    def as_dict(self) -> dict[str, Any]:
        """Header-keyed view for the `raw_row` jsonb column.

        Unnamed and duplicate headers get positional keys (`col_4`) so every
        cell survives the round trip — including the recruiter code hiding in
        the junk column.
        """
        out: dict[str, Any] = {}
        for index, name in enumerate(self.headers):
            key = _canonical(name) or f"col_{index + 1}"
            if key in out:
                key = f"col_{index + 1}"
            out[key] = self.cells[index] if index < len(self.cells) else None
        return out


@dataclass(frozen=True)
class ReadResult:
    sheet_name: str
    signature: HeaderSignature
    headers: tuple[str, ...]
    rows: list[RawRow]
    junk_columns: tuple[int, ...]
    """Indices of columns with no usable header — excluded from exports."""
    reported_max_row: int
    """What the sheet CLAIMED, kept so the phantom-row gap is observable."""
    rows_scanned: int


def read_rows(path: Path | str) -> ReadResult:
    """Read the first worksheet whose headers match a known signature."""
    path = Path(path)
    try:
        cached = load_workbook(path, read_only=True, data_only=True)
        formulas = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        # Covers a non-zip file, a corrupt archive, and anything else openpyxl
        # objects to. The filename is safe to surface; contents are not.
        raise NoDataSheetError(f"{path.name} could not be opened as a workbook") from exc

    try:
        return _read_matching_sheet(cached, formulas)
    finally:
        cached.close()
        formulas.close()


def _read_matching_sheet(cached: Any, formulas: Any) -> ReadResult:
    for sheet_name in cached.sheetnames:
        cached_sheet = cached[sheet_name]
        formula_sheet = formulas[sheet_name]

        found = _find_header_row(cached_sheet, formula_sheet)
        if found is None:
            continue
        header_index, headers, signature = found

        rows, scanned = _stream_rows(cached_sheet, formula_sheet, header_index, len(headers))
        return ReadResult(
            sheet_name=sheet_name,
            signature=signature,
            headers=tuple(headers),
            rows=[RawRow(number=n, cells=c, headers=tuple(headers)) for n, c in rows],
            junk_columns=tuple(i for i, h in enumerate(headers) if not _canonical(h)),
            reported_max_row=cached_sheet.max_row or 0,
            rows_scanned=scanned,
        )

    raise NoDataSheetError("no worksheet matched a known header signature")


def _find_header_row(cached_sheet: Any, formula_sheet: Any) -> tuple[int, list[str], Any] | None:
    """Locate the header row and identify the file, or None if this sheet is not it."""
    # strict=False: the two reads are of the same sheet, but a length
    # mismatch must degrade to a short read, never raise mid-parse.
    pairs = zip(
        cached_sheet.iter_rows(max_row=HEADER_SEARCH_DEPTH, values_only=True),
        formula_sheet.iter_rows(max_row=HEADER_SEARCH_DEPTH, values_only=True),
        strict=False,
    )
    for index, (cached_row, formula_row) in enumerate(pairs, start=1):
        headers = [_coalesce(c, f) for c, f in _pair_cells(cached_row, formula_row)]
        headers = ["" if h is None else str(h) for h in headers]
        signature = detect_signature(headers)
        if signature is not None:
            return index, headers, signature
    return None


def _stream_rows(
    cached_sheet: Any,
    formula_sheet: Any,
    header_index: int,
    width: int,
) -> tuple[list[tuple[int, tuple[Any, ...]]], int]:
    """Stream data rows, stopping at a run of blanks rather than at max_row."""
    rows: list[tuple[int, tuple[Any, ...]]] = []
    blank_run = 0
    scanned = 0

    pairs = zip(
        cached_sheet.iter_rows(min_row=header_index + 1, values_only=True),
        formula_sheet.iter_rows(min_row=header_index + 1, values_only=True),
        strict=False,
    )
    for offset, (cached_row, formula_row) in enumerate(pairs, start=header_index + 1):
        scanned += 1
        if scanned > MAX_ROWS_SCANNED:
            break

        cells = tuple(_coalesce(c, f) for c, f in _pair_cells(cached_row, formula_row))[:width]
        if all(_empty(cell) for cell in cells):
            blank_run += 1
            # A gap inside the data is normal; a long run means the data ended.
            if blank_run >= BLANK_RUN_LIMIT:
                break
            continue

        blank_run = 0
        rows.append((offset, cells))

    return rows, scanned


def _pair_cells(cached_row: tuple[Any, ...], formula_row: tuple[Any, ...]) -> Any:
    from itertools import zip_longest

    return zip_longest(cached_row or (), formula_row or (), fillvalue=None)


def _coalesce(cached: Any, formula: Any) -> Any:
    """Prefer the value Excel computed; fall back to the formula text.

    A formula Excel already evaluated arrives with its answer cached, and that
    answer is authoritative — including for `=H2*2.5`, which we could not work
    out ourselves. When there is no cached value we hand the normalizers the
    formula string to evaluate.
    """
    return formula if cached is None else cached


def _empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
