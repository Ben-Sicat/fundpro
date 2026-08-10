"""Authentication, charity scoping and PII containment.

These are the non-negotiables from CLAUDE.md and the PH Data Privacy Act
(RA 10173), so they get tested as behaviour rather than trusted as intent.
"""

from __future__ import annotations

import re

from fastapi.testclient import TestClient

from tests.conftest import TEST_API_KEY, ApiClient

PROTECTED = (
    "/pledges",
    "/kpis",
    "/uploads",
    "/exceptions",
    "/exports/templates",
    "/payroll/run",
    "/team/fundraisers",
    "/audit",
    "/settings/status-codes",
)


def test_every_endpoint_requires_a_bearer_token(client: TestClient) -> None:
    for path in PROTECTED:
        assert client.get(path).status_code == 401, path


def test_a_wrong_token_is_rejected(client: TestClient) -> None:
    response = client.get("/pledges", headers={"Authorization": "Bearer wrong-key"})
    assert response.status_code == 401


def test_a_malformed_authorization_header_is_rejected(client: TestClient) -> None:
    for header in ("", "Bearer", f"Basic {TEST_API_KEY}", TEST_API_KEY):
        response = client.get("/pledges", headers={"Authorization": header})
        assert response.status_code == 401, header


def test_health_stays_public_for_probes(client: TestClient) -> None:
    """No secret on the monitoring side; it leaks nothing but a status word."""
    assert client.get("/health").status_code in (200, 503)


def test_health_never_leaks_connection_details(client: TestClient) -> None:
    body = client.get("/health").text
    assert "secret-password" not in body
    assert "db.invalid" not in body


# ---------------------------------------------------------------------------
# charity_viewer scoping — enforced in the service layer, not the UI
# ---------------------------------------------------------------------------

SCOPED = {"X-Charity-Scope": "STC"}


def test_a_scoped_caller_sees_only_their_charity(loaded: ApiClient) -> None:
    rows = loaded.json("/pledges", headers=SCOPED)
    assert {r["charityCode"] for r in rows} == {"STC"}
    assert len(rows) == 3


def test_a_scoped_caller_cannot_widen_their_own_scope(loaded: ApiClient) -> None:
    """A query parameter must not be able to override the header scope."""
    rows = loaded.json("/pledges", params={"charity": "UNHCR"}, headers=SCOPED)
    assert rows == []


def test_a_scoped_caller_gets_a_404_not_a_403_for_another_charity(
    loaded: ApiClient,
) -> None:
    """A 403 would confirm the record exists. A 404 says nothing."""
    assert loaded.get("/pledges/FES48000003", headers=SCOPED).status_code == 404
    assert loaded.get("/pledges/FES48000001", headers=SCOPED).status_code == 200


def test_scoping_applies_to_the_dashboard_too(loaded: ApiClient) -> None:
    assert loaded.json("/kpis", headers=SCOPED)["signups"] == 3
    assert loaded.json("/charities", headers=SCOPED) == ["STC"]


def test_scoping_applies_to_events_and_notes(loaded: ApiClient) -> None:
    assert loaded.get("/pledges/FES48000003/events", headers=SCOPED).status_code == 404
    assert loaded.get("/pledges/FES48000003/notes", headers=SCOPED).status_code == 404


def test_a_scoped_caller_cannot_add_a_note_to_another_charity(loaded: ApiClient) -> None:
    response = loaded.post(
        "/pledges/FES48000003/notes", json={"text": "hello"}, headers=SCOPED
    )
    assert response.status_code == 404


def test_scoped_exports_contain_only_that_charity(loaded: ApiClient) -> None:
    from io import BytesIO

    from openpyxl import load_workbook

    response = loaded.post("/exports/A1", headers=SCOPED)
    ws = load_workbook(BytesIO(response.content), read_only=True).active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 3
    assert b"Carmela" not in response.content  # a UNHCR donor


# ---------------------------------------------------------------------------
# PII containment
# ---------------------------------------------------------------------------


def test_no_full_card_number_is_ever_returned(loaded: ApiClient) -> None:
    for path in ("/pledges", "/donors", "/pledges/FES48000001/events"):
        body = loaded.get(path).text
        assert not re.search(r"\d{13,19}", body), path


def test_import_exceptions_do_not_put_donor_names_in_the_detail(
    loaded: ApiClient,
) -> None:
    """The raw row is kept for an authorised human, but the message a
    developer reads in a log must be safe."""
    for exception in loaded.json("/exceptions"):
        assert "Alina" not in exception["detail"]
        assert "Someone Else Entirely" not in exception["detail"]


def test_the_audit_log_records_who_did_what(loaded: ApiClient) -> None:
    loaded.post("/exports/A1", headers={"X-Actor": "Rhea Santos"})
    entry = next(a for a in loaded.json("/audit") if a["action"] == "export.generate")
    assert entry["actor"] == "Rhea Santos"


def test_imports_are_audited(loaded: ApiClient) -> None:
    actions = {a["action"] for a in loaded.json("/audit")}
    assert "import.apps_tracker" in actions
    assert "import.status_report" in actions


def test_settings_changes_are_audited(loaded: ApiClient) -> None:
    loaded.put(
        "/settings/status-codes",
        json={"statusId": 90, "description": "Test", "classification": "other"},
    )
    assert "settings.status_code" in {a["action"] for a in loaded.json("/audit")}


def test_an_invalid_as_of_header_is_rejected_rather_than_ignored(
    loaded: ApiClient,
) -> None:
    """Silently falling back to today would make a wrong date look real."""
    response = loaded.get("/kpis", headers={"X-As-Of": "not-a-date"})
    assert response.status_code == 422
