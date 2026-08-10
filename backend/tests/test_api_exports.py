"""Exports, including the legacy masters that have to be byte-compatible."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.parsing.headers import APPS_TRACKER_COLUMNS, STATUS_REPORT_COLUMNS
from app.services.exports import A1_COLUMNS
from tests.conftest import ApiClient

XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def sheet(response):
    assert response.status_code == 200, response.text[:300]
    assert response.headers["content-type"] == XLSX_TYPE
    return load_workbook(BytesIO(response.content), read_only=True).active


def headers_of(ws) -> list[str]:
    return [c if c is not None else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]


# ---------------------------------------------------------------------------
# A1 — the one that has to match byte for byte
# ---------------------------------------------------------------------------


def test_a1_has_exactly_111_columns() -> None:
    """113 in the real file = 111 real + 2 junk."""
    assert len(A1_COLUMNS) == 111
    assert len(APPS_TRACKER_COLUMNS) == 113


def test_a1_drops_only_the_junk_columns() -> None:
    dropped = [h for h in APPS_TRACKER_COLUMNS if h not in A1_COLUMNS]
    # Position 4 is a single space, position 109 is empty.
    assert dropped == [" ", None]


def test_a1_preserves_the_header_quirks() -> None:
    """These oddities are load-bearing: the client's downstream tooling reads
    them literally, so 'fixing' the spelling breaks their process."""
    assert "CUSTOMER'S NAME" in A1_COLUMNS  # apostrophe
    assert "Fax AREACODE" in A1_COLUMNS  # lowercase 'F'
    assert "CHQ/MO/PO" in A1_COLUMNS
    assert "Invoice No." in A1_COLUMNS  # trailing period


def test_a1_keeps_the_original_column_order(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/A1"))
    assert headers_of(ws) == list(A1_COLUMNS)


def test_a1_round_trips_the_uploaded_applications(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/A1"))
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    serials = {row[A1_COLUMNS.index("SERIAL NO")] for row in rows}

    assert len(rows) == 6
    assert "FES48000001" in serials
    assert "FEH48000006" in serials


def test_a1_computes_age_at_export_time(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/A1"))
    age_index = A1_COLUMNS.index("AGE")
    serial_index = A1_COLUMNS.index("SERIAL NO")

    ages = {
        row[serial_index]: row[age_index]
        for row in ws.iter_rows(min_row=2, values_only=True)
    }
    # Born 1994-03-02; as of the fixed 2026-07-27 that is 32.
    assert ages["FES48000001"] == 32


# ---------------------------------------------------------------------------
# A2 / A3 — the bank schema
# ---------------------------------------------------------------------------


def test_a2_uses_the_26_bank_columns(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/A2"))
    assert headers_of(ws) == list(STATUS_REPORT_COLUMNS)
    assert len(STATUS_REPORT_COLUMNS) == 26


def test_a2_flattens_the_whole_billing_history(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/A2"))
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Seven events consolidated (three bank rows became exceptions).
    assert len(rows) == 7


def test_a3_adds_the_batch_columns_to_the_bank_schema(loaded: ApiClient) -> None:
    upload_id = loaded.json("/uploads")[0]["id"]
    ws = sheet(loaded.post("/exports/A3", params={"upload_id": upload_id}))
    assert headers_of(ws)[:26] == list(STATUS_REPORT_COLUMNS)
    assert headers_of(ws)[26:] == ["IMPORT BATCH ID", "IMPORTED AT"]


def test_a3_is_scoped_to_one_upload(loaded: ApiClient) -> None:
    status_upload = next(u for u in loaded.json("/uploads") if u["sourceType"] == "status_report")
    ws = sheet(loaded.post("/exports/A3", params={"upload_id": status_upload["id"]}))
    batch_ids = {row[26] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert batch_ids == {status_upload["id"]}


# ---------------------------------------------------------------------------
# Derived reports
# ---------------------------------------------------------------------------


def test_every_template_generates(loaded: ApiClient) -> None:
    for template in loaded.json("/exports/templates"):
        response = loaded.post(f"/exports/{template['code']}")
        assert response.status_code == 200, f"{template['code']} failed"
        assert len(response.content) > 0


def test_the_retry_queue_lists_only_failing_pledges(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/B2"))
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert {row[0] for row in rows} == {"FES48000003", "FEH48000004"}


def test_the_retry_queue_flags_an_expired_card(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/B2"))
    headers = headers_of(ws)
    expired_index = headers.index("CARD EXPIRED")
    rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
    # FEH48000004's card expired 01/26, before the fixed today of 2026-07-27.
    assert rows["FEH48000004"][expired_index] == "Y"


def test_the_verification_backlog_excludes_verified_donors(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/B3"))
    serials = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "FES48000001" not in serials  # verified
    assert "FEH48000006" in serials  # not called yet


def test_the_charity_delivery_report_contains_no_pii(loaded: ApiClient) -> None:
    """D2 is the one that gets emailed outward, so it must be safe by
    construction rather than by reviewer discipline."""
    response = loaded.post("/exports/D2")
    body = response.content

    for pii in (b"Alina", b"Bacani", b"example.invalid", b"542550"):
        assert pii not in body
    assert response.headers["x-contains-pii"] == "false"


def test_the_invoice_separates_charges_from_credits(loaded: ApiClient) -> None:
    ws = sheet(loaded.post("/exports/D1"))
    headers = headers_of(ws)
    kind = headers.index("LINE TYPE")
    lines = {row[0]: row[kind] for row in ws.iter_rows(min_row=2, values_only=True)}

    assert lines["FES48000001"] == "charge"
    # 0005 billed and then cancelled — the charity is credited back.
    assert lines["FES48000005"] == "credit"


# ---------------------------------------------------------------------------
# Catalogue, counts and the audit trail
# ---------------------------------------------------------------------------


def test_the_catalogue_reports_live_row_counts(loaded: ApiClient) -> None:
    templates = {t["code"]: t for t in loaded.json("/exports/templates")}
    assert templates["A1"]["rows"] == 6  # one per application
    assert templates["A2"]["rows"] == 7  # one per billing event
    assert templates["B4"]["rows"] == 3  # open exceptions


def test_counts_come_from_the_right_collection(loaded: ApiClient) -> None:
    """A wrong count is worse than none: 'Import Exceptions — 420 rows' sends
    someone hunting for a problem that does not exist."""
    templates = {t["code"]: t for t in loaded.json("/exports/templates")}
    assert templates["B4"]["rows"] != templates["A1"]["rows"]
    # Aggregates have no comparable row figure at all.
    assert templates["D2"]["rows"] is None


def test_row_counts_respect_the_active_filter(loaded: ApiClient) -> None:
    templates = {
        t["code"]: t for t in loaded.json("/exports/templates", params={"charity": "STC"})
    }
    assert templates["A1"]["rows"] == 3


def test_generating_an_export_is_audited_with_a_pii_flag(loaded: ApiClient) -> None:
    loaded.post("/exports/A1")
    loaded.post("/exports/D2")
    entries = [a for a in loaded.json("/audit") if a["action"] == "export.generate"]

    a1 = next(a for a in entries if "A1" in a["detail"])
    d2 = next(a for a in entries if "D2" in a["detail"])
    assert a1["containsPii"] is True
    assert d2["containsPii"] is False


def test_an_export_run_is_recorded(loaded: ApiClient) -> None:
    loaded.post("/exports/B1")
    runs = loaded.json("/exports/runs")
    assert runs[0]["templateCode"] == "B1"
    assert runs[0]["rowCount"] == 6
    assert runs[0]["fileName"].startswith("B1_")


def test_an_unknown_template_is_a_404(loaded: ApiClient) -> None:
    assert loaded.post("/exports/ZZ").status_code == 404


def test_the_filename_is_safe_for_a_content_disposition_header(loaded: ApiClient) -> None:
    response = loaded.post("/exports/D3")
    disposition = response.headers["content-disposition"]
    assert '"' in disposition
    # No stray quotes or newlines that would let a name break the header.
    assert "\n" not in disposition and disposition.count('"') == 2
