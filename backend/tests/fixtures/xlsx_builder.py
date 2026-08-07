"""Builds xlsx files that reproduce the traps in docs/FINDINGS.md §2.

Every generator here is deliberately hostile in the same ways the client's real
files are. If the parser survives these it survives production; if a new trap
turns up in the wild, add it here first and watch the test fail.

The headers come from `app.parsing.headers`, transcribed from the real
workbooks, so a fixture can never drift from the schema it claims to reproduce.

All donor-shaped values are synthetic. `.invalid` is reserved by RFC 2606 and
can never route. Masked card values follow the real files' `*` masking
(`542550********2906`), not the `X` form the older spec text assumed.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.parsing.headers import APPS_TRACKER_COLUMNS, STATUS_REPORT_COLUMNS

STATUS_REPORT_HEADERS: list[Any] = list(STATUS_REPORT_COLUMNS)
APPS_TRACKER_HEADERS: list[Any] = list(APPS_TRACKER_COLUMNS)

MASKED_PAN = "542550********2906"


def _save(wb: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _row(headers: list[Any], values: dict[str, Any]) -> list[Any]:
    """Build a positional row from a header→value map.

    Headers that repeat or are blank simply take None unless addressed by
    position through `values` using their 1-based index as a string key.
    """
    out: list[Any] = []
    for index, header in enumerate(headers, start=1):
        if str(index) in values:
            out.append(values[str(index)])
        elif header is not None and header in values:
            out.append(values[header])
        else:
            out.append(None)
    return out


def _status_row(serial: str, **over: Any) -> list[Any]:
    base: dict[str, Any] = {
        "Charity Code": "STC",
        "Bank": "HSBC",
        "SERIAL NO": serial,
        "STATUS ID": 66,
        "STATUS DESCRIPTION": "Billing Approved",
        "STATUS DATE": datetime(2026, 7, 8),
        "CUSTOMERS NAME": "Alina Bacani",
        "CREDIT CARD": MASKED_PAN,
        "A0 Attempts": 1,
        "Recruiter Batch No": "1783477131",
        "ExpiryDate": "1028",
        "DonationAmount": 1000,
        "Frequency": "1",
        "Recruiter Submission Date": datetime(2026, 7, 1),
        "AgentID": "FPH316",
        "DEBIT_CREDIT_CARD": "CREDIT",
        "Channel": "F2F",
        "Recruiter Code": "FP",
    }
    base.update(over)
    return _row(STATUS_REPORT_HEADERS, base)


def build_status_report(
    path: Path,
    *,
    sheet_name: str = "sheet1",
    phantom_rows: int = 0,
) -> Path:
    """A bank Status Report carrying the amount/date/expiry traps.

    `phantom_rows` appends fully-blank rows after the data, reproducing the
    sheets that report ~1,048,570 rows while holding a few hundred.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(STATUS_REPORT_HEADERS)

    rows = [
        # Trap 6: amount as comma-formatted TEXT.
        _status_row("FES48000001", **{"DonationAmount": "1,000.00"}),
        # Trap 4: amount as a literal formula string.
        _status_row("FES48000002", **{"DonationAmount": "=75*13"}),
        # Trap 5: zero-padded MMYY that dies if parsed as a number.
        _status_row("FES48000003", **{"ExpiryDate": "0728"}),
        # Trap 3: date as a literal formula string — as in the real file.
        _status_row("FES48000004", **{"STATUS DATE": "=DATE(2026,7,8)"}),
        # Trap 7: date as a plain 10-character string.
        _status_row("FES48000005", **{"STATUS DATE": "2026-07-08"}),
        # A failure row — status 59, the retryable one.
        _status_row(
            "FES48000006",
            **{
                "STATUS ID": 59,
                "STATUS DESCRIPTION": "Billing Failed (DNH - Will retry)",
                "REASON": "DO NOT HONOR",
                "REASONDESC": "DO NOT HONOR               82",
            },
        ),
        # A blank row INSIDE the data. Must be skipped without ending the read.
        [None] * len(STATUS_REPORT_HEADERS),
        _status_row("FES48000007"),
    ]
    for row in rows:
        ws.append(row)

    for _ in range(phantom_rows):
        ws.append([None] * len(STATUS_REPORT_HEADERS))

    return _save(wb, path)


def _apps_row(serial: str, **over: Any) -> list[Any]:
    base: dict[str, Any] = {
        "COUNTRY": "PH",
        "CHARITY CODE": "STC",
        "4": "FP",  # trap 8: junk header ' ', carries the recruiter code
        "SERIAL NO": serial,
        "CUSTOMER'S NAME": "Alina Bacani",
        "FIRSTNAME": "Alina",
        "LAST NAME": "Bacani",
        "GENDER": "FEMALE",
        "DOB": datetime(1994, 3, 2),
        "EMAIL": "alina.bacani1@example.invalid",
        "TEL HP": "9171234567",
        "CITY": "Cebu City",
        "CAMPAIGN CODE": "STC-F2F-2026",
        "PROCESSING BANK": "HSBC",
        "DONATION AMOUNT": 1000,
        "FREQUENCY": "Monthly",
        "CREDIT CARD": MASKED_PAN,
        "CARDTYPE": "CREDIT CARD",
        "EXPIRY": "1028",
        "ISSUING BANK": "HSBC Philippines",
        "EVENT CODE": "Mactan Cebu International Airport Terminal 2",
        "AGENT ID": "FPH316",
        "SIGNUP DATE": datetime(2026, 7, 1),
        "STATUS DATE": datetime(2026, 7, 3),
        "VERIFIEDDATE": "2026-07-05",
        "STATUS": "SUBMISSION",
        "Fundraiser Name": "Almara Pasco",
        "109": None,  # trap 8: the genuinely blank junk column
    }
    base.update(over)
    return _row(APPS_TRACKER_HEADERS, base)


def build_apps_tracker(path: Path, *, sheet_name: str = "Sheet2") -> Path:
    """An Apps Tracker carrying the junk-column and enum-drift traps."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(APPS_TRACKER_HEADERS)

    rows = [
        _apps_row("FES48000001"),
        # Trap 4: an unresolvable CELL-REFERENCE formula.
        _apps_row("FES48000002", **{"DONATION AMOUNT": "=H2*2.5"}),
        # Trap 4: resolvable literal arithmetic.
        _apps_row("FES48000003", **{"DONATION AMOUNT": "=75*13"}),
        # Trap 9: enum casing drift.
        _apps_row("FES48000004", **{"CARDTYPE": "Credit Card", "FREQUENCY": "Semi-annual"}),
        _apps_row("FES48000005", **{"CARDTYPE": "DEBIT CARD", "FREQUENCY": 1}),
        # Trap 10: free-text location with a trailing space.
        _apps_row("FES48000006", **{"EVENT CODE": "SM Light Mall "}),
    ]
    for row in rows:
        ws.append(row)
    return _save(wb, path)


def build_multi_sheet(path: Path) -> Path:
    """Workbook whose FIRST sheet is decorative and second holds the data.

    Trap 2: the real files use `sheet1`, `Sheet1` and `Sheet2` for equivalent
    content, so selection must go by header signature, never by name or index.
    """
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append(["Bank of Somewhere"])
    cover.append(["Daily status report — internal use only"])
    cover.append([])
    cover.append(["Generated", datetime(2026, 7, 8)])

    data = wb.create_sheet("Sheet1")
    data.append(STATUS_REPORT_HEADERS)
    data.append(
        _status_row(
            "FES48000010",
            **{
                "CUSTOMERS NAME": "Boyet Calderon",
                "ExpiryDate": "0728",
                "DonationAmount": "1,500.00",
                "STATUS DATE": "=DATE(2026,7,8)",
            },
        )
    )
    return _save(wb, path)


def build_inflated_dimensions(path: Path, *, real_rows: int = 5) -> Path:
    """Sheet whose reported dimensions vastly exceed its real content.

    Trap 1. Touching a far-away cell inflates the stored dimension record
    exactly as Excel does.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(STATUS_REPORT_HEADERS)
    for i in range(real_rows):
        ws.append(_status_row(f"FES4800{i:04d}", **{"CUSTOMERS NAME": "Carmela Dimaano"}))
    ws.cell(row=1_048_570, column=1).value = None
    return _save(wb, path)


def build_custom_status_report(path: Path, rows: list[dict[str, Any]]) -> Path:
    """A Status Report with caller-specified cell values, keyed by real header.

    For tests that need one deliberately broken cell. Keys are real column
    names, so a test can never drift from the actual schema.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(STATUS_REPORT_HEADERS)
    for values in rows:
        ws.append(_row(STATUS_REPORT_HEADERS, values))
    return _save(wb, path)


def build_empty_sheet(path: Path) -> Path:
    """Headers only, no data rows. Must parse to zero rows, not crash."""
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(STATUS_REPORT_HEADERS)
    return _save(wb, path)
