"""A small but complete two-file dataset for the API tests.

Six applications covering every billing outcome the business cares about, plus
bank rows that exercise each exception path. Deliberately hostile in the same
ways the real files are — comma amounts, formula dates, a zero-padded expiry —
so the API tests also prove the parser is wired in.

ALL VALUES ARE SYNTHETIC. `.invalid` is reserved by RFC 2606.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from app.parsing.headers import APPS_TRACKER_COLUMNS, STATUS_REPORT_COLUMNS

MASKED_PAN = {
    "FES48000001": "542550********2906",
    "FES48000002": "548809********3036",
    "FES48000003": "512345********7788",
    "FEH48000004": "545454********1212",
    "FES48000005": "542550********9090",
    "FEH48000006": "533333********4321",
}


def _row(headers: tuple[Any, ...], values: dict[str, Any]) -> list[Any]:
    out: list[Any] = []
    for index, header in enumerate(headers, start=1):
        if str(index) in values:
            out.append(values[str(index)])
        elif header is not None and header in values:
            out.append(values[header])
        else:
            out.append(None)
    return out


# ---------------------------------------------------------------------------
# Applications
# ---------------------------------------------------------------------------

APPLICATIONS: tuple[dict[str, Any], ...] = (
    {
        # Straightforward approval.
        "SERIAL NO": "FES48000001",
        "COUNTRY": "PH",
        "CHARITY CODE": "STC",
        "CUSTOMER'S NAME": "Alina Bacani",
        "EMAIL": "alina.bacani1@example.invalid",
        "TEL HP COUNTRYCODE": "+63",
        "TEL HP": "9171234567",
        "DOB": datetime(1994, 3, 2),
        "GENDER": "FEMALE",
        "CITY": "Cebu City",
        "DONATION AMOUNT": 1000,
        "FREQUENCY": "1",  # numeric code — exercises the frequency map
        "CARDTYPE": "CREDIT CARD",
        "EXPIRY": "1028",
        "ISSUING BANK": "HSBC Philippines",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 7, 1),
        "STATUS DATE": datetime(2026, 7, 3),
        "VERIFIEDDATE": "2026-07-05",
        "VERIFIED": "Y",
        "VERIFIEDBY": "Verification Desk",
        "Fundraiser Name": "Almara Pasco",
        "EVENT CODE": "MCIA T1 — July drive",
        "LOCATION CODE": "Mactan-Cebu International Airport Terminal 1",
        "AGENT ID": "FPH316",
        "STATUS": "SUBMISSION",
    },
    {
        # Rejected, then approved on retry — the owners' payroll question.
        "SERIAL NO": "FES48000002",
        "COUNTRY": "PH",
        "CHARITY CODE": "STC",
        "CUSTOMER'S NAME": "Boyet Calderon",
        "EMAIL": "boyet.calderon2@example.invalid",
        "TEL HP": "9179876543",
        "DOB": datetime(1988, 11, 9),
        "GENDER": "MALE",
        "CITY": "Mandaue",
        "DONATION AMOUNT": "=75*8",  # literal formula, resolves to 600
        "FREQUENCY": "Monthly",
        "CARDTYPE": "Debit Card",  # casing drift
        "EXPIRY": "0728",  # zero-padded: the leading zero is load-bearing
        "ISSUING BANK": "BDO Unibank",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 7, 1),
        "STATUS DATE": datetime(2026, 7, 2),
        "Fundraiser Name": "Almara Pasco",
        "EVENT CODE": "MCIA T1 — July drive",
        "AGENT ID": "FPH316",
        "STATUS": "SUBMISSION",
    },
    {
        # Still retrying.
        "SERIAL NO": "FES48000003",
        "COUNTRY": "PH",
        "CHARITY CODE": "UNHCR MY",  # alias — must canonicalise to UNHCR
        "CUSTOMER'S NAME": "Carmela Dimaano",
        "EMAIL": "carmela.dimaano3@example.invalid",
        "TEL HP": "9170001111",
        "DOB": datetime(1999, 6, 21),
        "GENDER": "FEMALE",
        "CITY": "Lapu-Lapu",
        "DONATION AMOUNT": "1,500.00",  # comma text
        "FREQUENCY": "3",
        "CARDTYPE": "CREDIT CARD",
        "EXPIRY": "0930",
        "ISSUING BANK": "Metrobank",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 7, 2),
        "STATUS DATE": "=DATE(2026,7,4)",  # formula date
        "Fundraiser Name": "Grace Tolentino",
        "EVENT CODE": "SM Light Mall — atrium ",  # trailing space
        "AGENT ID": "RC054",
        "STATUS": "SUBMISSION",
    },
    {
        # Failed for good — card expired.
        "SERIAL NO": "FEH48000004",
        "COUNTRY": "MALAYSIA",
        "CHARITY CODE": "WV",
        "CUSTOMER'S NAME": "Nabila Roslan",
        "EMAIL": "nabila.roslan4@example.invalid",
        "TEL HP": "123456789",
        "DOB": datetime(1975, 1, 30),
        "GENDER": "FEMALE",
        "CITY": "Kuala Lumpur",
        "DONATION AMOUNT": 800,
        "FREQUENCY": "12",
        "CARDTYPE": "credit",
        "EXPIRY": "0126",
        "ISSUING BANK": "Maybank Berhad",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 6, 20),
        "STATUS DATE": datetime(2026, 6, 22),
        "Fundraiser Name": "Grace Tolentino",
        "EVENT CODE": "LRT Sri Rampai — concourse",
        "AGENT ID": "FEH201",
        "STATUS": "SUBMISSION",
    },
    {
        # Approved, then cancelled — the clawback path. Already paid out.
        "SERIAL NO": "FES48000005",
        "COUNTRY": "PH",
        "CHARITY CODE": "STC",
        "CUSTOMER'S NAME": "Marisol Quiambao",
        "EMAIL": "marisol.quiambao5@example.invalid",
        "TEL HP": "9175554444",
        "DOB": datetime(1991, 9, 14),
        "GENDER": "FEMALE",
        "CITY": "Cebu City",
        "DONATION AMOUNT": 1200,
        "FREQUENCY": "Monthly",
        "CARDTYPE": "CREDIT CARD",
        "EXPIRY": "1129",
        "ISSUING BANK": "HSBC Philippines",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 7, 1),
        "STATUS DATE": datetime(2026, 7, 2),
        "VERIFIEDDATE": "2026-07-04",
        "VERIFIED": "Y",
        "Fundraiser Name": "Almara Pasco",
        "EVENT CODE": "MCIA T1 — July drive",
        "AGENT ID": "FPH316",
        "STATUS": "SUBMISSION",
        "Payout Date": datetime(2026, 7, 15),
    },
    {
        # Never submitted to the bank: pending, unverified.
        "SERIAL NO": "FEH48000006",
        "COUNTRY": "MALAYSIA",
        "CHARITY CODE": "WWF",
        "CUSTOMER'S NAME": "Chih Chien Hung",
        "EMAIL": "chihchien.hung6@example.invalid",
        "TEL HP": "198887777",
        "DOB": datetime(2001, 4, 4),
        "GENDER": "MALE",
        "CITY": "Petaling Jaya",
        "DONATION AMOUNT": 500,
        "FREQUENCY": "Monthly",
        "CARDTYPE": "DEBIT CARD",
        "EXPIRY": "0631",
        "ISSUING BANK": "Maybank Berhad",
        "PROCESSING BANK": "HSBC",
        "SIGNUP DATE": datetime(2026, 7, 20),
        "Fundraiser Name": "Grace Tolentino",
        "EVENT CODE": "Amcorp Mall — ground",
        "AGENT ID": "FEH201",
        "STATUS": "PENDING",
    },
)


def build_apps_workbook(path: Path, *, sheet_name: str = "Sheet2") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(APPS_TRACKER_COLUMNS))
    for record in APPLICATIONS:
        values = dict(record)
        values["4"] = "FP"  # the junk column that carries the recruiter code
        values["CREDIT CARD"] = MASKED_PAN[record["SERIAL NO"]]
        ws.append(_row(APPS_TRACKER_COLUMNS, values))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Bank outcomes
# ---------------------------------------------------------------------------


def _status(serial: str, **over: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "Charity Code": "STC",
        "Bank": "HSBC",
        "SERIAL NO": serial,
        "STATUS ID": 66,
        "STATUS DESCRIPTION": "Billing Approved",
        "STATUS DATE": datetime(2026, 7, 8),
        "CUSTOMERS NAME": "",
        "CREDIT CARD": MASKED_PAN.get(serial, ""),
        "A0 Attempts": 1,
        "Recruiter Batch No": "1783477131",
        "ExpiryDate": "1028",
        "DonationAmount": "1,000.00",
        "Frequency": "1",
        "Recruiter Submission Date": datetime(2026, 7, 3),
        "AgentID": "FPH316",
        "DEBIT_CREDIT_CARD": "CREDIT",
        "Channel": "F2F",
        "Recruiter Code": "FP",
    }
    base.update(over)
    return base


STATUS_ROWS: tuple[dict[str, Any], ...] = (
    _status("FES48000001", **{"CUSTOMERS NAME": "Alina Bacani"}),
    # Rejected then approved: two rows, two dates, two cutoffs.
    _status(
        "FES48000002",
        **{
            "STATUS ID": 59,
            "STATUS DESCRIPTION": "Billing Failed (DNH - Will retry)",
            "REASON": "DO NOT HONOR",
            "REASONDESC": "DO NOT HONOR               82",
            "STATUS DATE": datetime(2026, 7, 5),
            "CUSTOMERS NAME": "Boyet Calderon",
        },
    ),
    _status(
        "FES48000002",
        **{
            "STATUS DATE": "=DATE(2026,7,20)",  # formula date
            "A0 Attempts": 2,
            "CUSTOMERS NAME": "Boyet Calderon",
        },
    ),
    _status(
        "FES48000003",
        **{
            "STATUS ID": 59,
            "STATUS DESCRIPTION": "Billing Failed (DNH - Will retry)",
            "REASON": "DNH",
            "STATUS DATE": datetime(2026, 7, 9),
            "CUSTOMERS NAME": "Carmela Dimaano",
        },
    ),
    _status(
        "FEH48000004",
        **{
            "STATUS ID": 71,
            "STATUS DESCRIPTION": "Card Expired",
            "REASON": "EXPIRED",
            "STATUS DATE": datetime(2026, 6, 28),
            "CUSTOMERS NAME": "Nabila Roslan",
        },
    ),
    _status(
        "FES48000005",
        **{"STATUS DATE": datetime(2026, 7, 6), "CUSTOMERS NAME": "Marisol Quiambao"},
    ),
    _status(
        "FES48000005",
        **{
            "STATUS ID": 84,
            "STATUS DESCRIPTION": "Cancelled by Donor",
            "STATUS DATE": datetime(2026, 7, 22),
            "CUSTOMERS NAME": "Marisol Quiambao",
        },
    ),
    # --- rows that must become exceptions, not silent drops -----------------
    # A serial that is not in the applications master.
    _status("FES49999999", **{"CUSTOMERS NAME": "Nobody Here"}),
    # A status code the dictionary does not know.
    _status(
        "FES48000001",
        **{
            "STATUS ID": 77,
            "STATUS DESCRIPTION": "Something New",
            "STATUS DATE": datetime(2026, 7, 25),
            "CUSTOMERS NAME": "Alina Bacani",
        },
    ),
    # Right serial, wrong donor — the file is misaligned.
    _status(
        "FEH48000006",
        **{"CUSTOMERS NAME": "Someone Else Entirely", "STATUS DATE": datetime(2026, 7, 24)},
    ),
)


def build_status_workbook(path: Path, *, sheet_name: str = "sheet1") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(STATUS_REPORT_COLUMNS))
    for record in STATUS_ROWS:
        ws.append(_row(STATUS_REPORT_COLUMNS, record))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def build_status_subset(path: Path, serials: tuple[str, ...]) -> Path:
    """Just the rows for `serials` — for testing a second, partial upload."""
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(list(STATUS_REPORT_COLUMNS))
    for record in STATUS_ROWS:
        if record["SERIAL NO"] in serials:
            ws.append(_row(STATUS_REPORT_COLUMNS, record))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
