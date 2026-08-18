"""Custom exports: choose your own columns off the consolidated data."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from tests.conftest import ApiClient


def _sheet(payload: bytes):
    return load_workbook(io.BytesIO(payload), read_only=True).worksheets[0]


def test_the_field_catalogue_is_grouped(loaded: ApiClient) -> None:
    fields = loaded.json("/exports/fields", headers={"X-Allow-Pii": "true"})

    keys = {f["key"] for f in fields}
    assert {"serialNo", "donorName", "debitDate", "attemptsToSuccess"} <= keys
    assert {"Identity", "Donor", "Lifecycle", "Billing"} <= {f["group"] for f in fields}


def test_it_builds_a_sheet_with_exactly_the_chosen_columns(loaded: ApiClient) -> None:
    response = loaded.post(
        "/exports/custom/build",
        json={"columns": ["serialNo", "amount", "debitDate"], "name": "Cebu cut"},
    )

    assert response.status_code == 200, response.text
    sheet = _sheet(response.content)
    header = [c.value for c in next(sheet.iter_rows(max_row=1))]
    # The caller's order, not the catalogue's — they are building a sheet to
    # paste somewhere and the order is part of the request.
    assert header == ["Serial no", "Amount", "Debit date"]


def test_column_order_is_the_callers(loaded: ApiClient) -> None:
    response = loaded.post(
        "/exports/custom/build",
        json={"columns": ["debitDate", "serialNo"]},
    )

    header = [c.value for c in next(_sheet(response.content).iter_rows(max_row=1))]
    assert header == ["Debit date", "Serial no"]


def test_an_unknown_column_is_rejected(loaded: ApiClient) -> None:
    response = loaded.post(
        "/exports/custom/build", json={"columns": ["serialNo", "donorSalary"]}
    )

    assert response.status_code == 422
    assert "donorSalary" in response.text


def test_a_charity_viewer_is_not_offered_donor_contact_details(
    loaded: ApiClient,
) -> None:
    fields = loaded.json("/exports/fields", headers={"X-Charity-Scope": "STC"})

    keys = {f["key"] for f in fields}
    assert "donorEmail" not in keys
    assert "donorMobile" not in keys
    assert "maskedPan" not in keys
    # Non-PII columns are still available to them.
    assert "serialNo" in keys


def test_a_charity_viewer_cannot_export_pii_even_by_asking(loaded: ApiClient) -> None:
    """The scope rule is absolute and no header can re-enable it."""
    response = loaded.post(
        "/exports/custom/build",
        json={"columns": ["serialNo", "donorEmail"]},
        headers={"X-Charity-Scope": "STC", "X-Allow-Pii": "true"},
    )

    assert response.status_code == 403


def test_a_pii_export_is_flagged_in_the_audit_log(loaded: ApiClient) -> None:
    loaded.post(
        "/exports/custom/build",
        json={"columns": ["serialNo", "donorEmail"]},
        headers={"X-Allow-Pii": "true"},
    )

    entry = next(e for e in loaded.json("/audit") if e["action"] == "export.custom")
    assert entry["containsPii"] is True
    # WHICH columns left the building is what an auditor needs, and it names
    # no donor.
    assert "donorEmail" in entry["detail"]


def test_a_non_pii_export_is_not_flagged(loaded: ApiClient) -> None:
    loaded.post("/exports/custom/build", json={"columns": ["serialNo", "amount"]})

    entry = next(e for e in loaded.json("/audit") if e["action"] == "export.custom")
    assert entry["containsPii"] is False


def test_it_respects_the_current_filters(loaded: ApiClient) -> None:
    everything = loaded.post("/exports/custom/build", json={"columns": ["serialNo"]})
    scoped = loaded.post(
        "/exports/custom/build?charity=NOPE", json={"columns": ["serialNo"]}
    )

    assert int(scoped.headers["X-Row-Count"]) == 0
    assert int(everything.headers["X-Row-Count"]) > 0
